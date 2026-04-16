"""Genera Excel de avance del scraping desde SQLite."""
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from src.config import DATA_DIR

DB_PATH = DATA_DIR / "progress.db"
REPORT_PATH = DATA_DIR / f"AVANCE_SCRAPING_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_PROGRESO = PatternFill("solid", fgColor="FFEB9C")
FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")
FILL_PENDIENTE = PatternFill("solid", fgColor="D9E2F3")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(bold=True, size=14, color="1F4E79")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _estado_fill(estado: str) -> PatternFill:
    return {
        "completado": FILL_OK,
        "en_progreso": FILL_PROGRESO,
        "error": FILL_ERROR,
    }.get(estado, FILL_PENDIENTE)


def generar() -> Path:
    conn = _connect()
    wb = Workbook()

    # === Hoja 1: Resumen ===
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:F1")
    ws["A1"] = "AVANCE SCRAPING ONPE — ELECCIONES 2026"
    ws["A1"].font = FONT_TITLE

    ws["A3"] = "Generado:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"].font = Font(bold=True)

    # Totales
    resumen = conn.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN estado='completado' THEN 1 ELSE 0 END) as completados,
            SUM(CASE WHEN estado='en_progreso' THEN 1 ELSE 0 END) as en_progreso,
            SUM(CASE WHEN estado='error' THEN 1 ELSE 0 END) as errores,
            SUM(CASE WHEN estado='pendiente' THEN 1 ELSE 0 END) as pendientes,
            SUM(presidenciales) as total_actas,
            SUM(procesadas) as procesadas,
            SUM(pdfs_descargados) as pdfs
        FROM distritos
    """).fetchone()

    labels = [
        ("Distritos totales", resumen["total"]),
        ("Completados", resumen["completados"]),
        ("En progreso", resumen["en_progreso"]),
        ("Con error", resumen["errores"]),
        ("Pendientes", resumen["pendientes"]),
        ("", ""),
        ("Actas presidenciales", resumen["total_actas"] or 0),
        ("Actas procesadas", resumen["procesadas"] or 0),
        ("PDFs descargados", resumen["pdfs"] or 0),
    ]
    for i, (label, val) in enumerate(labels, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    # Progreso %
    total_actas = resumen["total_actas"] or 1
    procesadas = resumen["procesadas"] or 0
    ws.cell(row=15, column=1, value="Progreso actas").font = Font(bold=True, size=12)
    ws.cell(row=15, column=2, value=f"{procesadas}/{total_actas} ({procesadas*100/total_actas:.1f}%)").font = Font(size=12)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30

    # === Hoja 2: Distritos ===
    ws2 = wb.create_sheet("Distritos")
    headers = ["Distrito", "Ubigeo", "Estado", "Actas Total", "Presidenciales",
               "Procesadas", "% Avance", "Con Datos", "Sin PDF", "PDFs", "Error"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))

    distritos = conn.execute("""
        SELECT * FROM distritos ORDER BY
            CASE estado
                WHEN 'en_progreso' THEN 1
                WHEN 'error' THEN 2
                WHEN 'pendiente' THEN 3
                WHEN 'completado' THEN 4
            END, nombre
    """).fetchall()

    for row_idx, d in enumerate(distritos, start=2):
        pres = d["presidenciales"] or 1
        proc = d["procesadas"] or 0
        values = [
            d["nombre"], d["ubigeo"], d["estado"],
            d["total_actas"], d["presidenciales"], d["procesadas"],
            f"{proc*100/pres:.1f}%",
            d["con_datos"], d["sin_pdf"], d["pdfs_descargados"],
            d["error"] or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx == 3:
                cell.fill = _estado_fill(d["estado"])

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[chr(64 + col)].width = 18

    # === Hoja 3: Actas con error ===
    ws3 = wb.create_sheet("Errores")
    err_headers = ["Acta ID", "Mesa", "Distrito", "Error"]
    ws3.append(err_headers)
    _style_header(ws3, 1, len(err_headers))

    errores = conn.execute(
        "SELECT acta_id, mesa, distrito, error FROM actas WHERE error IS NOT NULL ORDER BY distrito"
    ).fetchall()

    for row_idx, e in enumerate(errores, start=2):
        for col_idx, val in enumerate([e["acta_id"], e["mesa"], e["distrito"], e["error"]], start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 60

    # === Hoja 4: PDFs pendientes ===
    ws4 = wb.create_sheet("PDFs Pendientes")
    pdf_headers = ["Mesa", "Distrito", "Tipo", "Archivo", "Error"]
    ws4.append(pdf_headers)
    _style_header(ws4, 1, len(pdf_headers))

    pdfs = conn.execute(
        "SELECT mesa, distrito, tipo, nombre_destino, error FROM pdfs WHERE descargado = 0 ORDER BY distrito"
    ).fetchall()

    tipo_map = {1: "ESCRUTINIO", 3: "INSTALACIÓN", 4: "SUFRAGIO"}
    for row_idx, p in enumerate(pdfs, start=2):
        vals = [p["mesa"], p["distrito"], tipo_map.get(p["tipo"], str(p["tipo"])),
                p["nombre_destino"], p["error"] or ""]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER

    conn.close()

    wb.save(str(REPORT_PATH))
    print(f"Reporte: {REPORT_PATH}")
    return REPORT_PATH


if __name__ == "__main__":
    generar()
