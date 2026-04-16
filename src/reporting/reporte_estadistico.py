"""Reporte para estadístico — Miraflores instalaciones + participación."""
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = "data/forensic.db"
OUT = "data/MIRAFLORES_estadistico.xlsx"


def _conn():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def sheet_instalaciones(conn) -> pd.DataFrame:
    """Sheet 1: todas las mesas con hora + participación + flags."""
    df = pd.read_sql_query("""
        SELECT
            i.mesa                          AS MESA,
            i.distrito                      AS DISTRITO,
            i.hora_instalacion_raw          AS HORA_INSTALACION,
            i.hora_instalacion_min          AS HORA_MIN,
            ROUND(i.hora_instalacion_min / 60.0, 2) AS HORA_DECIMAL,
            i.total_electores_habiles       AS ELECTORES_HABILES,
            a.total_votantes                AS TOTAL_VOTARON,
            ROUND(a.participacion_pct, 2)   AS PARTICIPACION_PCT,
            a.votos_validos                 AS VOTOS_VALIDOS,
            a.votos_blanco                  AS VOTOS_BLANCO,
            a.votos_nulos                   AS VOTOS_NULOS,
            i.material_buen_estado          AS MATERIAL_OK,
            i.observaciones                 AS OBSERVACIONES,
            a.estado_acta                   AS ESTADO_ACTA,
            CASE
                WHEN i.hora_instalacion_min < 420 THEN 'ANTES_7AM'
                WHEN i.hora_instalacion_min > 600 THEN 'DESPUES_10AM'
                WHEN i.hora_instalacion_min BETWEEN 420 AND 480 THEN 'NORMAL_7-8AM'
                ELSE 'NORMAL_8-10AM'
            END AS CATEGORIA_HORA,
            CASE WHEN i.hora_instalacion_min > 600 OR i.hora_instalacion_min < 420
                 THEN 1 ELSE 0 END AS FLAG_ANOMALIA_HORA,
            CASE WHEN i.observaciones IS NOT NULL THEN 1 ELSE 0 END AS FLAG_OBSERVACION
        FROM instalaciones i
        LEFT JOIN actas a ON a.mesa = i.mesa AND a.distrito = 'MIRAFLORES'
        WHERE i.error IS NULL AND i.hora_instalacion_raw IS NOT NULL
        ORDER BY i.hora_instalacion_min
    """, conn)
    return df


def sheet_sin_acta(conn) -> pd.DataFrame:
    """Sheet 2: mesas que votaron pero no tienen PDF de instalación."""
    df = pd.read_sql_query("""
        SELECT
            a.mesa                          AS MESA,
            a.distrito                      AS DISTRITO,
            a.total_electores               AS ELECTORES_HABILES,
            a.total_votantes                AS TOTAL_VOTARON,
            ROUND(a.participacion_pct, 2)   AS PARTICIPACION_PCT,
            a.votos_validos                 AS VOTOS_VALIDOS,
            a.estado_acta                   AS ESTADO_ACTA,
            'SIN_PDF_INSTALACION'           AS MOTIVO
        FROM actas a
        WHERE a.distrito = 'MIRAFLORES'
          AND a.mesa NOT IN (
              SELECT mesa FROM instalaciones
              WHERE error IS NULL AND hora_instalacion_raw IS NOT NULL
          )
        ORDER BY a.participacion_pct DESC NULLS LAST
    """, conn)
    return df


def sheet_observaciones(conn) -> pd.DataFrame:
    """Sheet 3: solo mesas con observaciones textuales — evidencia forense."""
    df = pd.read_sql_query("""
        SELECT
            i.mesa                          AS MESA,
            i.hora_instalacion_raw          AS HORA_INSTALACION,
            i.hora_instalacion_min          AS HORA_MIN,
            i.total_electores_habiles       AS ELECTORES_HABILES,
            a.total_votantes                AS TOTAL_VOTARON,
            ROUND(a.participacion_pct, 2)   AS PARTICIPACION_PCT,
            i.observaciones                 AS OBSERVACION_LITERAL,
            CASE
                WHEN i.observaciones LIKE '%cedula%marcada%' OR i.observaciones LIKE '%marcad%'
                     THEN 'CEDULA_MARCADA'
                WHEN i.observaciones LIKE '%cedula%arrugada%' OR i.observaciones LIKE '%doblada%'
                     THEN 'CEDULA_DANADA'
                WHEN i.observaciones LIKE '%304%' OR i.observaciones LIKE '%extra%'
                     THEN 'CEDULAS_EXTRA'
                WHEN i.observaciones LIKE '%impresora%' OR i.observaciones LIKE '%manual%'
                     THEN 'FALLA_SISTEMA'
                WHEN i.observaciones LIKE '%suplente%' OR i.observaciones LIKE '%secretario%'
                     THEN 'IRREGULARIDAD_MESA'
                WHEN i.observaciones LIKE '%hora%' OR i.observaciones LIKE '%10am%' OR i.observaciones LIKE '%sistema%'
                     THEN 'HORA_MANIPULADA'
                ELSE 'OTRA'
            END AS TIPO_IRREGULARIDAD
        FROM instalaciones i
        LEFT JOIN actas a ON a.mesa = i.mesa AND a.distrito = 'MIRAFLORES'
        WHERE i.observaciones IS NOT NULL
        ORDER BY TIPO_IRREGULARIDAD, i.hora_instalacion_min
    """, conn)
    return df


def _formato_excel(ws, df: pd.DataFrame, titulo: str, color_header: str = "1F4E79"):
    """Aplica formato profesional a una hoja."""
    # Título
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=color_header)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")

    ws["A2"] = f"Distrito: MIRAFLORES  |  Total registros: {len(df)}  |  Fuente: ONPE + PDF vía AI"
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.row_dimensions[2].height = 14

    # Header
    header_fill = PatternFill("solid", fgColor=color_header)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30

    # Datos con color condicional
    flag_cols = {c: i+1 for i, c in enumerate(df.columns)
                 if "FLAG" in c or "ANOMALIA" in c or "CATEGORIA" in c or "TIPO" in c}
    hora_col = next((i+1 for i, c in enumerate(df.columns) if c == "HORA_MIN"), None)
    part_col = next((i+1 for i, c in enumerate(df.columns) if c == "PARTICIPACION_PCT"), None)

    fill_red   = PatternFill("solid", fgColor="FFD7D7")
    fill_amber = PatternFill("solid", fgColor="FFF3CD")
    fill_green = PatternFill("solid", fgColor="D6F5D6")
    fill_alt   = PatternFill("solid", fgColor="F7F9FC")

    for row_idx, row in enumerate(df.itertuples(index=False), 4):
        row_fill = fill_alt if row_idx % 2 == 0 else None
        hora_min = getattr(row, "HORA_MIN", None) if hasattr(row, "HORA_MIN") else None
        flag_anomalia = getattr(row, "FLAG_ANOMALIA_HORA", 0) if hasattr(row, "FLAG_ANOMALIA_HORA") else 0

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(df.columns)))
            cell.font = Font(size=9)

            # Color por anomalía
            if flag_anomalia and hora_col and col_idx == hora_col:
                cell.fill = fill_red
            elif col_idx == part_col and value is not None:
                if value >= 90:
                    cell.fill = fill_amber
                elif row_fill:
                    cell.fill = row_fill
            elif row_fill:
                cell.fill = row_fill

    # Anchos de columna
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df.iloc[:, col_idx-1].astype(str).str.len().max() or 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    # Freeze + filtros
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(df.columns))}3"


def generar(output: str = OUT) -> str:
    conn = _conn()
    df1 = sheet_instalaciones(conn)
    df2 = sheet_sin_acta(conn)
    df3 = sheet_observaciones(conn)
    conn.close()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="1_INSTALACIONES", index=False, startrow=2)
        df2.to_excel(writer, sheet_name="2_SIN_ACTA_INSTALACION", index=False, startrow=2)
        df3.to_excel(writer, sheet_name="3_OBSERVACIONES_FORENSES", index=False, startrow=2)

    # Aplicar formato
    from openpyxl import load_workbook
    wb = load_workbook(output)

    _formato_excel(wb["1_INSTALACIONES"], df1,
        "MIRAFLORES — Horas de Instalación de Mesas (418 mesas)", "1F4E79")
    _formato_excel(wb["2_SIN_ACTA_INSTALACION"], df2,
        "MIRAFLORES — Mesas SIN Acta de Instalación (62 mesas)", "7B2D2D")
    _formato_excel(wb["3_OBSERVACIONES_FORENSES"], df3,
        "MIRAFLORES — Observaciones Forenses en Actas", "2D5A27")

    wb.save(output)

    print(f"\nReporte: {output}")
    print(f"  Sheet 1 — Instalaciones:       {len(df1)} mesas")
    print(f"    Anomalias hora:              {df1['FLAG_ANOMALIA_HORA'].sum()}")
    print(f"    Con observaciones:           {df1['FLAG_OBSERVACION'].sum()}")
    print(f"  Sheet 2 — Sin acta:            {len(df2)} mesas")
    print(f"  Sheet 3 — Obs. forenses:       {len(df3)} registros")
    return output


if __name__ == "__main__":
    generar()
