"""Anexos probatorios por distrito para EXPEDIENTES_JNE.

Genera por cada distrito:
  02_ANEXO_MESAS.xlsx        - listado foliado de mesas con hora, votos, estado
  03_CATALOGO_HASHES.pdf     - SHA-256 de cada PDF probatorio
  04_CADENA_CUSTODIA.csv     - export del log de custodia del distrito
  05_INDICE_PIEZA_PROBATORIA.pdf - indice del expediente

Cliente final omitido por confidencialidad (rol perito imparcial).
"""
from __future__ import annotations

import csv
import hashlib
import logging
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak,
)

DB_PATH = Path("data/forensic.db")
OUT_BASE = Path("data/EXPEDIENTES_JNE")

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

DISTRITOS = [
    "MIRAFLORES", "SAN JUAN DE MIRAFLORES", "SANTIAGO DE SURCO",
    "VILLA EL SALVADOR", "SAN ISIDRO", "PUCUSANA", "MAGDALENA DEL MAR",
]


def _slug(d: str) -> str:
    return d.replace(" ", "_")


def _conn() -> sqlite3.Connection:
    return sqlite3.connect(DB_PATH)


# -------------------- 02 ANEXO MESAS XLSX --------------------
def anexo_mesas_xlsx(distrito: str, out_path: Path) -> int:
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            a.mesa,
            a.local_votacion,
            i.hora_instalacion_raw,
            i.hora_instalacion_min,
            a.total_electores,
            a.total_votantes,
            a.votos_emitidos,
            a.votos_validos,
            a.votos_blanco,
            a.votos_nulos,
            a.estado_acta,
            a.flag_sin_acta
        FROM actas a
        LEFT JOIN instalaciones i ON i.acta_id = a.acta_id
        WHERE a.distrito = ?
        ORDER BY a.mesa
    """, (distrito,))
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mesas"

    header_fill = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    head = [
        "FOLIO", "MESA", "LOCAL_VOTACION", "HORA_INST_RAW", "HORA_INST_MIN",
        "TOT_ELECT", "TOT_VOT", "EMITIDOS", "VALIDOS", "BLANCO", "NULOS",
        "ESTADO_ACTA", "SIN_ACTA",
    ]
    for col, h in enumerate(head, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for idx, r in enumerate(rows, start=1):
        ws.cell(row=idx + 1, column=1, value=f"F-{idx:05d}")
        for j, v in enumerate(r, start=2):
            ws.cell(row=idx + 1, column=j, value=v)

    widths = [10, 14, 40, 14, 12, 11, 11, 11, 11, 9, 9, 16, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Metadatos")
    meta_rows = [
        ("Distrito", distrito),
        ("Total mesas", len(rows)),
        ("Generado (UTC)", datetime.utcnow().isoformat() + "Z"),
        ("Foliacion", f"F-00001 a F-{len(rows):05d}"),
        ("Fuente", "data/forensic.db (SQLite v2 forense)"),
        ("Tablas", "actas LEFT JOIN instalaciones ON acta_id"),
        ("Pieza", "Anexo 02 — Listado foliado de mesas"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        meta.cell(row=i, column=1, value=k).font = Font(bold=True)
        meta.cell(row=i, column=2, value=str(v))
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 60

    wb.save(out_path)
    return len(rows)


# -------------------- 03 CATALOGO HASHES PDF --------------------
def catalogo_hashes_pdf(distrito: str, out_path: Path) -> int:
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            mesa,
            CASE tipo WHEN 1 THEN 'ESCRUTINIO'
                     WHEN 3 THEN 'INSTALACION'
                     WHEN 4 THEN 'SUFRAGIO'
                     ELSE 'OTRO' END,
            nombre_destino,
            tamano_bytes,
            sha256_hash,
            descarga_at
        FROM pdfs
        WHERE distrito = ? AND descargado = 1 AND sha256_hash IS NOT NULL
        ORDER BY mesa, tipo
    """, (distrito,))
    rows = cur.fetchall()
    conn.close()

    doc = SimpleDocTemplate(
        str(out_path), pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title=f"Catalogo de hashes - {distrito}",
    )
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph(f"<b>ANEXO 03 - CATALOGO DE HASHES SHA-256</b>", styles["Title"]))
    elems.append(Paragraph(f"Distrito: <b>{distrito}</b>", styles["Normal"]))
    elems.append(Paragraph(
        f"Generado (UTC): {datetime.utcnow().isoformat()}Z &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Total piezas: <b>{len(rows)}</b>",
        styles["Normal"],
    ))
    elems.append(Spacer(1, 0.5 * cm))
    elems.append(Paragraph(
        "Cada hash SHA-256 fue calculado al momento de descarga del archivo "
        "desde la URL S3 firmada del servidor ONPE. Garantiza la integridad "
        "binaria de la pieza probatoria. Cualquier alteracion posterior "
        "produce un hash distinto.",
        styles["Italic"],
    ))
    elems.append(Spacer(1, 0.4 * cm))

    head = ["Folio", "Mesa", "Tipo", "Bytes", "SHA-256", "Descarga (UTC)"]
    data = [head]
    for i, r in enumerate(rows, start=1):
        mesa, tipo, _nombre, size, sha, fecha = r
        sha_short = (sha[:16] + "..." + sha[-8:]) if sha and len(sha) > 28 else (sha or "-")
        data.append([
            f"H-{i:05d}", str(mesa), tipo, f"{size or 0:,}", sha_short,
            (fecha or "")[:19],
        ])

    t = Table(data, colWidths=[1.4 * cm, 2.0 * cm, 2.4 * cm, 2.0 * cm, 6.5 * cm, 3.5 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2A44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]))
    elems.append(t)

    elems.append(Spacer(1, 0.6 * cm))
    elems.append(Paragraph(
        "<b>Nota tecnica:</b> los hashes completos (64 caracteres hex) se "
        "encuentran en la tabla <i>pdfs</i> de la base de datos forense "
        "<i>data/forensic.db</i> y son verificables ejecutando el script "
        "<i>src/audit/integrity.py</i>.",
        styles["Italic"],
    ))

    doc.build(elems)
    return len(rows)


# -------------------- 04 CADENA CUSTODIA CSV --------------------
def cadena_custodia_csv(distrito: str, out_path: Path) -> int:
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, timestamp, accion, entidad_tipo, entidad_id,
               operador, maquina, version_software, detalle
        FROM cadena_custodia
        WHERE detalle LIKE ? OR entidad_id LIKE ?
        ORDER BY timestamp
    """, (f"%{distrito}%", f"%{distrito}%"))
    rows = cur.fetchall()
    conn.close()

    head = [
        "id", "timestamp", "accion", "entidad_tipo", "entidad_id",
        "operador", "maquina", "version_software", "detalle",
    ]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(head)
        w.writerows(rows)
    return len(rows)


# -------------------- 05 INDICE PIEZA PROBATORIA --------------------
def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def indice_pieza_probatoria_pdf(distrito: str, expediente_dir: Path) -> None:
    out_path = expediente_dir / f"05_INDICE_PIEZA_PROBATORIA_{_slug(distrito)}.pdf"

    piezas = []
    for f in sorted(expediente_dir.iterdir()):
        if f.name.startswith("05_INDICE"):
            continue
        if not f.is_file():
            continue
        size_kb = f.stat().st_size / 1024
        sha = _file_sha256(f)
        piezas.append((f.name, f"{size_kb:,.1f} KB", sha))

    doc = SimpleDocTemplate(
        str(out_path), pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"Indice pieza probatoria - {distrito}",
    )
    styles = getSampleStyleSheet()
    elems = []

    elems.append(Paragraph("<b>INDICE DE PIEZA PROBATORIA</b>", styles["Title"]))
    elems.append(Paragraph(
        f"Expediente de nulidad - Distrito de <b>{distrito}</b>",
        styles["Heading2"],
    ))
    elems.append(Paragraph(
        f"Generado (UTC): {datetime.utcnow().isoformat()}Z",
        styles["Normal"],
    ))
    elems.append(Spacer(1, 0.5 * cm))

    elems.append(Paragraph(
        "El presente indice enumera la totalidad de piezas integrantes del "
        "expediente probatorio remitido para sustentar el pedido de nulidad. "
        "Cada pieza se acompana de su hash SHA-256 a fin de permitir su "
        "verificacion individual.",
        styles["Normal"],
    ))
    elems.append(Spacer(1, 0.5 * cm))

    head = ["#", "Pieza", "Tamano", "SHA-256"]
    data = [head]
    for i, (name, size, sha) in enumerate(piezas, start=1):
        sha_short = sha[:24] + "..." + sha[-8:]
        data.append([str(i), name, size, sha_short])

    t = Table(data, colWidths=[1 * cm, 8.5 * cm, 2.4 * cm, 6 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2A44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]))
    elems.append(t)

    elems.append(Spacer(1, 0.8 * cm))
    elems.append(Paragraph(
        "<b>Constancia de integridad:</b> los hashes consignados fueron "
        "generados por el algoritmo SHA-256 (FIPS 180-4) ejecutado sobre el "
        "contenido binario completo de cada pieza al momento de la firma de "
        "este indice. Cualquier modificacion ulterior producira un hash "
        "distinto.",
        styles["Italic"],
    ))
    elems.append(Spacer(1, 1 * cm))

    elems.append(Paragraph(
        f"Documento confidencial. Referencia: {distrito} - "
        f"{datetime.utcnow().strftime('%Y-%m-%d')}.",
        styles["Italic"],
    ))

    doc.build(elems)


# -------------------- ORQUESTADOR --------------------
def generar_anexos_distrito(distrito: str) -> dict:
    expediente_dir = OUT_BASE / f"EXPEDIENTE_NULIDAD_{_slug(distrito)}"
    expediente_dir.mkdir(parents=True, exist_ok=True)

    slug = _slug(distrito)

    f02 = expediente_dir / f"02_ANEXO_MESAS_{slug}.xlsx"
    f03 = expediente_dir / f"03_CATALOGO_HASHES_{slug}.pdf"
    f04 = expediente_dir / f"04_CADENA_CUSTODIA_{slug}.csv"

    n_mesas = anexo_mesas_xlsx(distrito, f02)
    n_pdfs = catalogo_hashes_pdf(distrito, f03)
    n_log = cadena_custodia_csv(distrito, f04)
    indice_pieza_probatoria_pdf(distrito, expediente_dir)

    return {
        "distrito": distrito,
        "mesas": n_mesas,
        "pdfs": n_pdfs,
        "log_custodia": n_log,
        "expediente_dir": str(expediente_dir),
    }


def main() -> None:
    log.info(f"{'DISTRITO':25}{'MESAS':>8}{'PDFS':>8}{'LOG':>8}")
    log.info("-" * 50)
    for d in DISTRITOS:
        try:
            r = generar_anexos_distrito(d)
            log.info(f"{r['distrito']:25}{r['mesas']:>8}{r['pdfs']:>8}{r['log_custodia']:>8}")
        except Exception as e:
            log.error(f"FALLO {d}: {e}")


if __name__ == "__main__":
    main()
